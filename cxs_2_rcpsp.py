# -*- coding: utf-8 -*-
"""CXS_2_RCPSP.ipynb

Automatically generated by Colaboratory.

Original file is located at
    https://colab.research.google.com/drive/1ceVjUqw1N_Z2YSW2rBUJOJt3lSkWfsgw

Complex Scheduling, Summer term 2023

*Prof. Dr. Rainer Kolisch, M.Sc. Pia Ammann*

# Assignment 2: RCPSP

**[25+2 points]**

In this assignment, you should implement different RCPSP formulations and compare their computational performance.

**Instructions**:
Please submit your solution together with the names and matriculation numbers of all team members until Sunday, 06-25-2023 (23:59) via email to pia.ammann@tum.de.
Use "*CXS23 Assignment 2 [last name student 1 / last name student 2 (/last name student 3 )]*" for the subject line.

**Team members**: name student 1 / name student 2 (/ name student 3 )

## 0 Basics

This section introduces the basic data structures and sets up Gurobi. Please read the descriptions carefully and execute all code cells to make use of the defined variables.
In the last part (performance improvement), you should complete the code snipet (copy paste from assignment 1).

### Data Structures (extended)

An object of the class `Network` represents a network of activities with the (dummy) starting node $1$, the (dummy) finishing node $n$ and $n-2$ activity nodes. All precedence relations and timelags are stored in the attribute arcs (dictionary).

An object of class `Node` represents a node with an index and attributes earliest start time (est), latest start time (lst), earliest finish time (eft), latest finish time (eft), and start time (start). Furthermore, every node has a set of successors and a processing time.

Note that the data structures slightly changed compared to assignment 1! (E.g., a `Node` object now has an additional attribute `resource_requirements`, which stores the resource demand for every resource (in a dictonary)).
"""

# data structures
class Node:

    def __init__(self, index, successors, processing_time, resource_requirements):
        self.id=index
        self.successors=successors
        self.processing_time=processing_time
        self.resource_requirements=resource_requirements
        self.est=None
        self.lst=None
        self.eft=None
        self.lft=None
        self.start=None


    def __str__(self):
        return 'Node {} with p={}, r={}'.format(self.id, self.processing_time,
                                                self.resource_requirements)



class Network:

    def __init__(self, name, node_dict, arcs, k, C):
        self.name = name
        self.node_ids = list(range(0,len(node_dict.keys()))) # I changed the range from range(1,len(node_dict.keys())+1)
        self.node_dict = node_dict
        self.arcs=arcs                  # dict of timelags (i,j): d_ij
        self.Tmax=sum([n.processing_time for n in self.node_dict.values()])+1
        self.makespan=None
        # additional attributes for RCPSP
        self.number_of_resources = k
        self.resource_availability = C  # dict of resource availabilities


    def __str__(self):
        return '{}: A network with {}+2 nodes, {} arcs, and deadline {}.'.format(self.name, len(self.node_ids)-2, len(self.arcs), self.Tmax)

"""### Gurobi Setup

As in assignment 1, make sure to install Gurobi and create an environment with your WLS license.

(https://support.gurobi.com/hc/en-us/articles/4409582394769-Google-Colab-Installation-and-Licensing)
"""

# mount drive to access data files
from google.colab import drive
drive.mount('/content/drive')


import gurobipy as gp
from gurobipy import GRB
from copy import deepcopy

# 1 Read license from file

lic_path = 'D:\Van\Complex Scheduling\gurobi.lic'

llic = {}
lic = 2372825

# 2 Create environment with your license
e = gp.Env(empty=True)
e.setParam('LicenseID', lic)
e.start()

# 2 Create environment with your license
e = gp.Env(empty=True)
for param, value in lic.items():
  e.setParam(param, value)
e.start()

"""### Performance Improvement

In the *discrete-time* formulations of the RCPSP, the start times of activities are modeled by using binary decision variables $x_{it}$ or $y_{it}$ with $x_{it}, y_{it}\in\{0,1\}$ for all $t \in \{ES_i,\ldots,LS_i\}$. The smaller the set $\{ES_i,\ldots,LS_i\}$, the fewer decision variables have to be considered and RCPSP instances can be solved faster.

To improve the performance of your MIP implementations, you can decrease the cardinality of set $\{ES_i,\ldots,LS_i\}$, which will decrease the number of time-discrete variables. However, you must not exclude any possible optimal start times of activities: As a naive approach, you can set $ES_i=0$ and $LS_i=\sum_{i=0}^{n+1}d_i$. You will get smaller sets though, if you set $E_i$ to the earliest permitted start times of the Generic Time Scheduling Problem (without resource limitations). These start times can be calculated by either using the *Label-Correcting-Algorithm*, the *Floyd-Warshall-Algorithm*, or a formulation of the *generic time scheduling problem* (with the objective function $min \sum_{i=0}^{n+1}S_i$).

Please choose one of the three options to reduce the time window size, and complete the method `compute_time_windows()` accordingly. This method takes a `Network` instance as input and sets the attributes `est`, `lst` of all `Node` objects accordingly.
(You can simply copy-paste your code from Assignment 1!)
"""

def compute_time_windows(network):

  def lca(network):
    # init
    d = [0] + [float('-inf') for n in network.node_ids[1:]] # list for distances
    p = [0] + [-1 for n in network.node_ids[1:]]            # list for predecessors
    q = [0]                                              # queue

    #main
    while len(q) > 0:
      i = q.pop(0)
      for j in network.node_dict[i].successors:
        #print(network.arcs)
        #print(network.node_dict[0].successors)
        #print(network.arcs)
        #print(j, i)
        #print(network.node_dict[i].successors)
        #print(j, i, d[j], d[i], network.arcs[(i, j)])
        if d[j] < d[i] + network.arcs[(i, j)]:
          d[j] = d[i] + network.arcs[(i, j)]
          p[j] = i
          if d[j] > ((len(network.node_ids)-1) * max(network.arcs.values())):
            return "Positive Cycles, algorithm stopped."
          if j not in q:
            q.append(j)
    return d, p, q

  def create_auxiliary_network(network):

    auxiliary_network = Network(network.name, deepcopy(network.node_dict), deepcopy(network.arcs), network.number_of_resources, network.resource_availability)
    auxiliary_network.arcs[((len(auxiliary_network.node_ids) - 1), 0)] = -network.Tmax
    # print(auxiliary_network.arcs)
    # print(len(network.node_ids))
    keys = list(auxiliary_network.arcs.keys())
    # print(network.arcs.keys())
    values = list(auxiliary_network.arcs.values())
    # print(network.arcs.values())
    tuple_rev = [(sub[1], sub[0]) for sub in keys]
    # print(tuple_rev)
    arc_rev = {tuple_rev[i]: values[i] for i in range(0, len(auxiliary_network.arcs))}
    # print(arc_rev)
    auxiliary_network.arcs = arc_rev
    # successors have to be corrected
    #print(auxiliary_network.node_dict[0])
    for n in auxiliary_network.node_ids:
      auxiliary_network.node_dict[n].successors =[i for i in auxiliary_network.node_ids if (n,i) in auxiliary_network.arcs]
    #print(auxiliary_network.node_dict[4].successors
    #print(network.node_dict)
    #print(auxiliary_network.node_dict)
    # print(auxiliary_network.arcs)
    return auxiliary_network


  # calculateing EST / LST
  d, p, q = lca(network)
  #print(d)
  for k, v in network.node_dict.items():
    v.est = d[k]
    v.eft = d[k] + network.node_dict[k].processing_time

  auxiliary_network = create_auxiliary_network(network)
  s, t, u = lca(auxiliary_network)
  #print(auxiliary_network.node_dict[2].est)
  #print(d)
  for k, v in network.node_dict.items():
    #print(d)
    v.lst = -s[k]
    v.lft = v.lst + network.node_dict[k].processing_time
    #print(network.node_dict[k].lst)
    #print(network.node_dict[k].lft)

  return

"""## 1 RCPSP Formulations

Given a network $N=(V,E, \delta_{ij})$ with activities $V$, precedence relations $E$, and time lag relations $\delta_{ij}$, the *Resource-Constrained Project Scheduling Problem (RCPSP)*  holds as follows:

\begin{aligned}
&\text{Minimize } f(\mathbf{S})&\\
&s.t.\\
&S_j-S_i\geq \delta_{ij} & \forall (i,j)\in E\\
&S_0=0&\\
&r_k(S,t) \leq R_k & \forall k\in R, t \in [0, T^{max})
\end{aligned}

Remarks:
* Resource constraint is due to $r_k(S,t)$ non-linear and non-convex.
* This model is conceptual and cannot be implemented as LP or MIP.
* This problem is a generalization of the Job Shop-Problems and thus NP-hard.

In chapter 2 of the "Handbook on Project Management and Scheduling Vol. 1", Artigues et al. (2015) present different possibilities to formulate the *Resource-Constrained Project Scheduling Problem (RCPSP)* as a mixed integer linear optimization problem. (Chapter 2 of the book is available on Moodle.)

\

In the following, you should **implement different formulations for the RCPSP**. For every formulation, you should complete the respective `solve()` method, which formulates and solves an optimization model of the RCPSP with the objective to minimize the makespan for a given network. Use the solver Gurobi.

These methods should ***not*** print the default solver output but instead print...
* the name of the ***test record*** together with the ***formulation*** used,
* the optimal ***objective value***,
* and the ***runtime*** of the solver in seconds.

If the runtime for a run exceeds ***10 minutes***, stop the computation and document the best value found until then. For not optimally solved instances, also report the percentage deviation (***gap***) between the best found solution and the best bound.

Furthermore, these methods should **set** the nodes attributes (`start`) and network attributes (`makespan`) accordingly.

Finally, every `solve()` method should **return** the following information:

* best found (/optimal) objective value
* runtime of the solver in seconds
* MIP gap

(You will need this information later, when you have to write your results to Excel)

### 1.1 Pulse Variables
Implement the ***Discrete-Time Formulations Based on "Pulse" Start Variables*** (see chapter 2.2.1) in the *aggregated* (DT) and the **disaggregated** (DDT) version.

a.) Aggregated (DT)

**[3.5 points]**
"""

from time import process_time
def solve_dt_pulse(network):

  # Creating the model within Gurobi
  m = gp.Model('PSP', env=e)

  # Run Parameters
  m.Params.LogToConsole = 0
  m.Params.timelimit = 30

  #defining parameters and variables
  V = len(network.node_ids)
  E = list(network.arcs.keys())
  H = int(network.Tmax)
  C = network.resource_availability
  x = m.addVars(V, H, vtype=GRB.BINARY, name='pulse variable')

  # Constraints

  for (i, j) in E:
    ct_2_2_lhs = gp.LinExpr()
    for t in range(0, H):
      ct_2_2_lhs += t * x[j, t]
      ct_2_2_lhs -= t* x[i, t]
    m.addConstr(ct_2_2_lhs >= network.node_dict[i].processing_time, name='ct_2_2')

  for t in range(0, H):
    for k in range(0, len(C)):
      ct_2_3_lhs = gp.LinExpr()
      for i in range(0, V):
        for tau in range(t-network.node_dict[i].processing_time + 1, t+1):
          if tau >= 0:
            ct_2_3_lhs += network.node_dict[i].resource_requirements[k] * x[i, tau]
      m.addConstr(ct_2_3_lhs <= C[k], name='ct_2_3')

  for i in range(0, V):
    ct_2_4_lhs = gp.LinExpr()
    for t in range(0, H):
      ct_2_4_lhs += x[i, t]
    m.addConstr(ct_2_4_lhs >= 1, name='ct_2_4a')
    m.addConstr(ct_2_4_lhs <= 1, name='ct_2_4b')

  for i in range(0, V):
    tw = set(range(network.node_dict[i].est, network.node_dict[i].lst+1))
    for t in range(0, H):
      if t not in tw:
        m.addConstr(x[i, t] >= 0, name='ct_2_5a')
        m.addConstr(x[i, t] <= 0, name='ct_2_5b')

  # ct 2_6 (x is binary) included in var definition

  # OF
  m.modelSense = gp.GRB.MINIMIZE
  objective = gp.LinExpr()

  objective = 0
  for t in range(0, H):
      objective += t * x[V-1 , t] # V correct? debug

  m.setObjective(objective)
  m.optimize()

  # set makespan and node starting times
  network.makespan = objective
  for i in network.node_ids:
    network.node_dict[i].start = 0
    for t in range(0, H):
      network.node_dict[i].start += x[i, t]*t

  if m.status == GRB.OPTIMAL :
    # Output
    print("With the dt_pulse a solution was found for " + str(network.name))
    print("The best found objective value is: " + str(m.ObjVal))
    print("Solver runtime: " + str(m.runtime))
    gap = m.Params.mipgap
    ov = m.getObjective()
    rt = m.runtime

    # Additional Output for Testing
    #for t in range(0, H):
      #for i in range(0, V):
        #print(f"x_{i, t} = {x[i, t].X}")

    #for v in m.getVars():
      #print(f"{v.VarName} = {v.X}")


    return ov, rt, gap

  else:
    print("Optimal solution not found")
    print("With the dt_pulse a solution was found for " + str(network.name))
    print("The best found objective value is: " + str(m.ObjVal))
    print("Solver runtime: " + str(m.runtime))
    gap = m.Params.mipgap
    ov = m.getObjective()
    rt = m.runtime
    return

"""b.) Disaggregated (DDT)

**[1.5 points]**
"""

from os import name
def solve_ddt_pulse(network):

  # Creating the model within Gurobi
  m = gp.Model('PSP', env=e)

  # Run Parameters
  m.Params.LogToConsole = 0
  m.Params.timelimit = 30 # change later to 600

  #defining parameters and variables
  V = len(network.node_ids)
  E = list(network.arcs.keys())
  H = int(network.Tmax)
  C = network.resource_availability
  x = m.addVars(V, H, vtype=GRB.BINARY, name='pulse variable')

  # Constraints

  for (i, j) in E:
      for t in range(0, H):
        ct_2_7_lhs = gp.LinExpr()
        for tau in range(0, t-network.node_dict[i].processing_time+1): # range correct?
          ct_2_7_lhs += x[i, tau]
        for tau in range(0, t+1):
          ct_2_7_lhs -= x[j, tau]
        m.addConstr(ct_2_7_lhs >= 0, name='ct_2_7')

  for t in range(0, H):
    for k in range(0, len(C)):
      ct_2_3_lhs = gp.LinExpr()
      for i in range(0, V):
        for tau in range(t-network.node_dict[i].processing_time + 1, t+1):
          if tau >= 0:
            ct_2_3_lhs += network.node_dict[i].resource_requirements[k] * x[i, tau]
      m.addConstr(ct_2_3_lhs <= C[k], name='ct_2_3')

  for i in range(0, V):
    ct_2_4_lhs = gp.LinExpr()
    for t in range(0, H):
      ct_2_4_lhs += x[i, t]
    m.addConstr(ct_2_4_lhs >= 1, name='ct_2_4a')
    m.addConstr(ct_2_4_lhs <= 1, name='ct_2_4b')

  for i in range(0, V):
    tw = set(range(network.node_dict[i].est, network.node_dict[i].lst+1))
    for t in range(0, H):
      if t not in tw:
        m.addConstr(x[i, t] >= 0, name='ct_2_5a')
        m.addConstr(x[i, t] <= 0, name='ct_2_5b')

  # ct 2_6 (x is binary) included in var definition

  # OF
  m.modelSense = gp.GRB.MINIMIZE
  objective = gp.LinExpr()

  objective = 0
  for t in range(0, H):
      objective += t * x[V-1 , t]

  m.setObjective(objective)
  m.optimize()

  # set makespan and node starting times
  network.makespan = objective
  for i in network.node_ids:
    network.node_dict[i].start = 0
    for t in range(0, H):
      network.node_dict[i].start += x[i, t]*t

  if m.status == GRB.OPTIMAL :
    # Output
    print("With the ddt_pulse a solution was found for " + str(network.name))
    print("The best found objective value is: " + str(m.ObjVal))
    print("Solver runtime: " + str(m.runtime))
    gap = m.Params.mipgap
    ov = m.getObjective()
    rt = m.runtime

    # Additional Output for Testing
    #for t in range(0, H):
      #for i in range(0, V):
        #print(f"x_{i, t} = {x[i, t].X}")

    #for v in m.getVars():
      #print(f"{v.VarName} = {v.X}")


    return ov, rt, gap

  else:
    print("Optimal solution not found")
    print("With the ddt_pulse a solution was found for " + str(network.name))
    print("The best found objective value is: " + str(m.ObjVal))
    print("Solver runtime: " + str(m.runtime))
    gap = m.Params.mipgap
    ov = m.getObjective()
    rt = m.runtime
    return ov, rt, gap

"""### 1.2 Step Variables
Implement the *Discrete-Time Formulations Based on "Step" Start Variables* (see chapter 2.2.2) in the aggregated (DT) and the disaggregated (DDT) version.

a.) Aggregated (DT)

**[3.5 points]**
"""

def solve_dt_step(network):

  # Creating the model within Gurobi
  m = gp.Model('PSP', env=e)

  # Run Parameters
  m.Params.LogToConsole = 0
  m.Params.timelimit = 60 # change later to 660

  #defining parameters and variables
  V = len(network.node_ids)
  E = list(network.arcs.keys())
  H = int(network.Tmax)
  C = network.resource_availability
  eta = m.addVars(V, H+1, vtype=GRB.BINARY, name='step variable')

  # constraints

  for (i, j) in E:
    ct_2_16_lhs = gp.LinExpr()
    for t in range(1, H):
      ct_2_16_lhs += t * (eta[j, t] - eta[j, t-1])
      ct_2_16_lhs -= t * (eta[i, t] - eta[i, t-1])
    m.addConstr(ct_2_16_lhs >= network.node_dict[i].processing_time, name='ct_2_16')

    for t in range(0, H):
      for k in range(0, len(C)):
        ct_2_11_lhs = gp.LinExpr()
        for i in range(0, V):
          if t-network.node_dict[i].processing_time >= 0:
            ct_2_11_lhs += network.node_dict[i].resource_requirements[k] * (eta[i, t] - eta[i, t-network.node_dict[i].processing_time])
            m.addConstr(ct_2_11_lhs <= network.resource_availability[k], name='ct_2_11')

    for i in range(0, V):
      ct_2_12_lhs = gp.LinExpr()
      ct_2_12_lhs += eta[i, network.node_dict[i].lst] # lst index correct?
      m.addConstr(ct_2_12_lhs >= 1, name='ct_2_12a')
      m.addConstr(ct_2_12_lhs <= 1, name='ct_2_12b')

    for i in range(0, V):
      for t in range(1, H): # I start from 1 to avoid undefined index "t = -1"
          ct_2_13_lhs = gp.LinExpr()
          ct_2_13_lhs += eta[i, t] - eta[i, t-1]
          m.addConstr(ct_2_13_lhs >= 0, name='ct_2_13')

    for i in range (0, V):
      for t in range(0, H):
        if t <= network.node_dict[i].est-1:
          m.addConstr(eta[i, t] >= 0, name='ct_2_14a')
          m.addConstr(eta[i, t] <= 0, name='ct_2_14b')

    # ct15 implicitly stated in var definition ?

  # OF
  m.modelSense = gp.GRB.MINIMIZE
  objective = gp.LinExpr()

  objective = 0
  for t in range(1, H):
      objective += t * (eta[V-1, t] - eta[V-1, t-1]) # V correct?

  m.setObjective(objective)
  m.optimize()

  # set makespan and node starting times
  network.makespan = objective
  for i in range(0, V):
    network.node_dict[i].start = 0
    for t in range(1, H):
      network.node_dict[i].start = t * (eta[i, t] - eta[i, t-1])


  if m.status == GRB.OPTIMAL:
    # Output
    print("With the dt_step a solution was found for " + str(network.name))
    print("The best found objective value is: " + str(m.ObjVal))
    print("Solver runtime: " + str(m.runtime))
    gap = m.Params.mipgap
    ov = m.getObjective()
    rt = m.runtime

    # Additional Output for Testing
    #for t in range(0, H):
      #for i in range(0, V):
        #print(f"eta_{i, t} = {eta[i, t].x}")

    return ov, rt, gap

  else:
    print("Optimal solution not found")
    print("With the dt_step a solution was found for " + str(network.name))
    print("The best found objective value is: " + str(m.ObjVal))
    print("Solver runtime: " + str(m.runtime))
    gap = m.Params.mipgap
    ov = m.getObjective()
    rt = m.runtime

    return

"""b.) Disaggregated (DDT)

**[1.5 points]**
"""

def solve_ddt_step(network):

  # Creating the model within Gurobi
  m = gp.Model('PSP', env=e)

  # Run Parameters
  m.Params.LogToConsole = 0
  m.Params.timelimit = 90 # change later to 660

  #defining parameters and variables
  V = len(network.node_ids)
  E = list(network.arcs.keys())
  H = int(network.Tmax)
  C = network.resource_availability
  eta = m.addVars(V, H+1, vtype=GRB.BINARY, name='step variable')

  # constraints

  for (i, j) in E:
    for t in range(0, H):
      if t-network.node_dict[i].processing_time >= 0:
        ct_2_10_lhs = gp.LinExpr()
        ct_2_10_lhs += eta[i, t-network.node_dict[i].processing_time] - eta[j, t]
        m.addConstr(ct_2_10_lhs >= 0, name='ct_2_10')

  for t in range(0, H):
    for k in range(0, len(C)):
      ct_2_11_lhs = gp.LinExpr()
      for i in range(0, V):
        if t-network.node_dict[i].processing_time >= 0:
          ct_2_11_lhs += network.node_dict[i].resource_requirements[k] * (eta[i, t] - eta[i, t-network.node_dict[i].processing_time])
          m.addConstr(ct_2_11_lhs <= network.resource_availability[k], name='ct_2_11')

  for i in range(0, V):
    ct_2_12_lhs = gp.LinExpr()
    ct_2_12_lhs += eta[i, network.node_dict[i].lst] # lst index correct?
    m.addConstr(ct_2_12_lhs >= 1, name='ct_2_12a')
    m.addConstr(ct_2_12_lhs <= 1, name='ct_2_12b')

  for i in range(0, V):
    for t in range(1, H): # I start from 1 to avoid undefined index "t = -1"
        ct_2_13_lhs = gp.LinExpr()
        ct_2_13_lhs += eta[i, t] - eta[i, t-1]
        m.addConstr(ct_2_13_lhs >= 0, name='ct_2_13')

  for i in range (0, V):
    for t in range(0, H):
      if t <= network.node_dict[i].est-1:
        m.addConstr(eta[i, t] >= 0, name='ct_2_14a')
        m.addConstr(eta[i, t] <= 0, name='ct_2_14b')

  # ct15 implicitly stated in var definition ?

  # OF
  m.modelSense = gp.GRB.MINIMIZE
  objective = gp.LinExpr()

  objective = 0
  for t in range(1, H):
      objective += t * (eta[V-1, t] - eta[V-1, t-1]) # V correct?

  m.setObjective(objective)
  m.optimize()

  # set makespan and node starting times
  network.makespan = objective
  for i in range(0, V):
    network.node_dict[i].start = 0
    for t in range(1, H):
      network.node_dict[i].start = t * (eta[i, t] - eta[i, t-1])


  if m.status == GRB.OPTIMAL:
    # Output
    print("With the ddt_step a solution was found for " + str(network.name))
    print("The best found objective value is: " + str(m.ObjVal))
    print("Solver runtime: " + str(m.runtime))
    gap = m.Params.mipgap
    ov = m.getObjective()
    rt = m.runtime

    # Additional Output for Testing
    #for t in range(0, H):
      #for i in range(0, V):
        #print(f"eta_{i, t} = {eta[i, t].x}")

    return ov, rt, gap

  else:
    print("Optimal solution not found")
    print("With the ddt_step a solution was found for " + str(network.name))
    print("The best found objective value is: " + str(m.ObjVal))
    print("Solver runtime: " + str(m.runtime))
    gap = m.Params.mipgap
    ov = m.getObjective()
    rt = m.runtime

    return

"""### 1.3 Flow Variables

Implement the flow-based formulation with continuous-time variables as introduced in the lecture.

**[5 points]**
"""

def solve_flow(network):

  # Creating the model within Gurobi
  m = gp.Model('PSP', env=e)

  # Run Parameters
  m.Params.LogToConsole = 0
  m.Params.timelimit = 120 # change later to 600

  #defining parameters and variables
  V = len(network.node_ids)
  E = list(network.arcs.keys())
  C = network.resource_availability
  S = m.addVars(V, lb = 0, vtype=GRB.CONTINUOUS, name='starting time variable')
  f = m.addVars(V, V, C, lb  = 0, vtype=GRB.CONTINUOUS, name='resource flow variable')
  x = m.addVars(range(0, V-1), range(1, V), vtype=GRB.BINARY, name='activity sequencing variable')

  # constraints
  for (i, j) in E:
    m.addConstr(x[i, j] >= 1, name='ct_2_17a')
    m.addConstr(x[i, j] <= 1, name='ct_2_17b')

  for i in range(0, V-1):
    for j in range(1, V):
      ct_2_18_lhs = gp.LinExpr()
      ct_2_18_lhs += S[j] - S[i] - (2*network.Tmax) * x[i, j]
      ct_2_18_rhs = gp.LinExpr()
      ct_2_18_rhs += network.node_dict[i].processing_time - (2*network.Tmax)
      m.addConstr(ct_2_18_lhs >= ct_2_18_rhs, name="ct_2_18")

  for i in range(0, V-1):
    for j in range(1, V):
      for k in range(0, len(C)):
        ct_2_19_lhs = gp.LinExpr()
        if network.node_dict[i].resource_requirements[k] <= network.node_dict[j].resource_requirements[k]:
          ct_2_19_lhs += f[i, j, k] - network.node_dict[i].resource_requirements[k] * x[i, j]
        else:
          ct_2_19_lhs += f[i, j, k] - network.node_dict[j].resource_requirements[k] * x[i, j]
        m.addConstr(ct_2_19_lhs <= 0, name="ct_2_19")

  for i in range(0, V-1):
    for k in range(0, len(C)):
      ct_2_20_lhs = gp.LinExpr()
      for j in range(1, V):
        ct_2_20_lhs += f[i, j, k]
      m.addConstr(ct_2_20_lhs >= network.node_dict[i].resource_requirements[k], name="ct_2_20a")
      m.addConstr(ct_2_20_lhs <= network.node_dict[i].resource_requirements[k], name="ct_2_20b")

  for j in range(1, V):
    for k in range(0, len(C)):
      ct_2_21_lhs = gp.LinExpr()
      for i in range(0, V-1):
        ct_2_21_lhs += f[i, j, k]
      m.addConstr(ct_2_21_lhs >= network.node_dict[j].resource_requirements[k], name="ct_2_21a")
      m.addConstr(ct_2_21_lhs <= network.node_dict[j].resource_requirements[k], name="ct_2_21b")

  #for i in range(0, V-1):
    #for j in range(1, V):
      #if i == j:
        #m.addConstr(x[i, j] >= 0, name="ct_2_22a")
        #m.addConstr(x[i, j] <= 0, name="ct_2_22b")

  for i in range(0, V-1):
    for j in range(1, V):
      for k in range(0, len(C)):
        #m.addConstr(f[i, j, k] >= 0, name="ct_2_23a")
        if i == j:
          m.addConstr(f[i, j, k] >= 0, name="ct_2_23b") #<=


  m.addConstr(S[0] >= 0, name="ct_2_24a")
  m.addConstr(S[0] <= 0, name="ct_2_24b")

  # OF
  m.modelSense = gp.GRB.MINIMIZE
  objective = gp.LinExpr()

  objective = S[V-1]

  m.setObjective(objective)
  m.optimize()

  # set makespan and node starting times
  network.makespan = objective
  for i in range(1, V):
    network.node_dict[i].start = S[i]

  if m.status == GRB.OPTIMAL:
    # Output
    for v in m.getVars():
      print(f"{v.VarName} = {v.X}")
      print("test")
    print("With the flow model a solution was found for " + str(network.name))
    print("The best found objective value is: " + str(m.ObjVal))
    print("Solver runtime: " + str(m.runtime))
    gap = m.Params.mipgap
    ov = m.getObjective()
    rt = m.runtime


    # Additional Output for Testing


    return ov, rt, gap

  else:
    print(S[2])
    print("test")
    print("Optimal solution not found")
    print("With the flow model a solution was found for " + str(network.name))
    print("The best found objective value is: " + str(m.ObjVal))
    print("Solver runtime: " + str(m.runtime))
    gap = m.Params.mipgap
    ov = m.getObjective()
    rt = m.runtime

    return ov, rt, gap

"""## 2 Computational Study
Now, you can finally test your implementations! 🙃

Test your implementation with the test data sets provided as an Excel document (file $\texttt{RCPSP_data.xlsx}$). Each test instance contains 30 real activities with their resource requirements, durations and direct successors. Test instances are taken from the [PSPLIB instance library](http://www.om-db.wi.tum.de/psplib/) [(Kolisch and Sprecher, 1997)](https://www.sciencedirect.com/science/article/pii/S0377221796001701).

File format of the provided input file $\texttt{RCPSP_data.xlsx}$:

$n \qquad k\\
R_1 \qquad ... \qquad R_k\\
p_1 \qquad r_{1,1} \qquad ... \qquad r_{1,k} \qquad s_1 \qquad j^{1}_{1} \qquad ... \qquad j^{1}_{s_1}\\
... \qquad ... \qquad ... \qquad ... \qquad ... \qquad ... \quad ... \qquad ...\\
p_{n-1} \quad r_{n-1,1} \quad ... \qquad r_{n-1,k} \quad s_{n-1} \quad j^{n-1}_{1} \quad ... \qquad j^{n-1}_{s_n-1}\\
p_n \qquad r_{n,1} \qquad ... \qquad  r_{n,k} \qquad s_n \qquad j^{n}_{1} \qquad ... \qquad j^{n}_{s_n}\\
$

\

Symbol || Explanation
---|---|---
$n$ || Number of activities (incl. dummy start/end)
$k$ || Number of (renewable) resources
$R_k$ || Availability of resource $k$
$p_i$ || Processing time of activity $i \in \{1, ..., n\}$
$r_{ik}$ || Requirement of resource $k$ for activity $i$
$j^{i}_{s}$ || Activity index of the s-th successor of activity $i$

### 2.1 Read Data

**[2 points]**
"""

# read data from excel using pandas

from openpyxl import load_workbook
import pandas as pd

xlsx_link = 'D:\Van\Complex Scheduling\Copy of RCPSP_data.xlsx'
workbook = load_workbook(xlsx_link)
Instances = {}

# iterate through all sheets (instances)
for sheet in workbook.worksheets:

    # get general info first
    sheet_name = sheet.title

for sheet in workbook.worksheets[1:3]:

  sheet_name = sheet.title

  n = int(sheet.cell(2, 1).value)

  k = int(sheet.cell(2, 2).value)

  C = dict() # dict with resources (index from 0 to k-1 as key, availability as values)
  for i in range(0, k):
    C[i] = sheet.cell(3, i+1).value

  S_n = []
  for nn in range(0, n):
    S_n.append(int(sheet.cell(4+nn, 6).value))

  node_dict = {i: Node(index=n, successors=[], processing_time=0, resource_requirements=[] ) for i in range(0, n)}

  for nn in range(0, n):
    for kk in range(0, k):
      node_dict[nn].resource_requirements.append(sheet.cell(nn+4, kk+2).value)

  for nn in range(0, n):
    node_dict[nn].processing_time = int(sheet.cell(4+nn, 1).value)

  for nn in range(0, n):
    for s in range(0, S_n[nn]):
      node_dict[nn].successors.append(int(sheet.cell(nn+4, s+7).value-1)) # -1 bc node notation from 0 to 31

  time_lags = dict()
  for i in range(0, n):
    for j in range(0, len(node_dict[i].successors)):
      time_lags[(i, node_dict[i].successors[j])] = int(node_dict[i].processing_time)

  Instances[sheet_name] = Network(name=sheet_name,
                                node_dict=node_dict,
                                arcs=time_lags,
                                k=k, C=C)

"""### 2.2 Results

Write a loop, which **solves all test records** stored in Instances one after another with **all five formulations**, making use of the respective `solve()` methods defined in part 1. For all time-discrete formulations, make sure to *tighten the time windows* making use of the previously defined method `compute_time_windows()`.

Additionally, your code should **write** the key information (best objective, runtime, MIP gap) to the "Results" tab in the provided Excel file $\texttt{RCPSP_data.xlsx}$.

Make sure to also fill in your hardware specifications and to compute the "Gap to optimality [%]" as well as the number of instances solved to optimality (row 57) for every formulation (you might want to do this directly in Excel).


**[3 points]**
"""

for inst in Instances:
  network = Instances[inst]

  compute_time_windows(network)

  solve_dt_pulse(network)
  #solve_ddt_pulse(network)
  #solve_dt_step(network)
  #solve_ddt_step(network)
  #solve_flow(network)

# write results to results tab

# fill in hardware specs
#workbook.worksheets.title

# gap to optimality

# nr of instances solved to optimality

"""### 2.3 Comparison

Finally, you should compare the performance of the different MIP formulations by means of a small computational analysis, as you would typically do in a scientific paper.

When comparing different solution approaches / formulations, you are usually interested in the following questions:
* How / in which respects does the performance differ among the approaches?
* Which approach performs best w.r.t. runtime / bounds / objective value?
* What are the avg, med, min, max deviation from best known values?
* ...


To answer these questions, you should conduct a small analysis computing and comparing the following values:

* $solved [\%]$: Share of instances for which a feasible solution was found
* $opt.solved [\%]$: Share of instances solved to proven optimality
* $t [s]$: Average runtime **to optimality**
* $\Delta \; z [\%]$: Average deviation from best known objective value
* $\Delta \; gap [\%]$: Average deviation from best known MIP gap

You can conduct your analysis directly in Excel or using python.

\

Report your results by filling in the **table** below...

Measure / sol approach | P-DT | P-DDT | S-DT | S-DDT | F-CT
---|---|---|---|---|---
$solved [\%]$ | | | | |
$opt.solved [\%]$ | | | | |
$t [s]$ | | | | |
$\Delta \; z [\%]$ | | | | |
$\Delta \; gap [\%]$ | | | | |

..and summarize your **key findings** in 2-3 sentences:

`# your answer here`

\

**[5 points]**
"""

# insert your code for part 2.3 here (if needed):
#
#
#

"""### 2.4 Visualization (optional!)

The table above shows only average values. To provide additional information, e.g., on distributions, min / max values, or median values, you can use boxplots showing the distritubtion of your KPIs for every formulation. It's sufficient to create plots for one of the KPIs reported in the Table. This gives 2 extra points.

**[2 points]**
"""

# insert your code for part 2.4 here:
#
#
#